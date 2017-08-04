import simplejson as json
import glob
import sys
import os
from openpyxl import load_workbook, Workbook
import numpy as np

def main():

	wb = load_workbook("processedGamesKDA.xlsx")
	wb_t = load_workbook("processedGamesTestKDA.xlsx")
	wb_s = load_workbook("processedGamesTestKDA.xlsx")


	players = [66271229, 44207669, 47063254, 56723622, 34399881, 47837396, 19305039, 37348109, 51635691, 28360391]
	normalizeTest(players,wb,wb_s, wb_t)
	# normalize(players,wb,wb_s)
	# fixKDA(players, wb)
	# merge(players)
	# fixTestKDA()

def merge(players):

	wb = load_workbook("dataSetKDA.xlsx")
	wb2 = load_workbook("normalizedKDA.xlsx")
	ws2 = wb2["training"]

	letters = ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O", "P", "Q", "R", "S", "T", "U", "V", "W", "X"]
	i = 2
	for playerId in players:

		print(playerId)

		ws = wb[str(playerId)]

		j = 2
		while ws["A" + str(j)].value!=None:

			for letter in letters:

				ws2[letter + str(i)] = ws[letter + str(j)].value

			j += 1
			i += 1

	wb2.save("normalizedKDA.xlsx")

def fixTestKDA():

	wb = load_workbook("processedGamesTest.xlsx")
	ws = wb["39550290"]

	i = 2

	while ws["A" + str(i)].value!=None:

		if ws["G" + str(i)].value != None:

			if ws["I" + str(i)].value == 0:
				ws["AC" + str(i)] = ws["G" + str(i)].value + ws["H" + str(i)].value
			else:
				ws["AC" + str(i)] = float(ws["G" + str(i)].value + ws["H" + str(i)].value)/(ws["I" + str(i)].value)

		if ws["T" + str(i)].value != None:

			if ws["T" + str(i)].value == 0:
				ws["AD" + str(i)] = ws["R" + str(i)].value + ws["S" + str(i)].value
			else:
				ws["AD" + str(i)] = float(ws["R" + str(i)].value + ws["S" + str(i)].value)/ws["T" + str(i)].value

		i += 1

	wb.save("processedGamesTestKDA.xlsx")


def fixKDA(players,wb):

	wb = load_workbook("processedGames_correct.xlsx")
	for playerId in players:

		ws = wb[str(playerId)]

		i = 2

		while ws["A" + str(i)].value!=None:

			if ws["G" + str(i)].value != None:

				if ws["I" + str(i)].value == 0:
					ws["AC" + str(i)] = ws["G" + str(i)].value + ws["H" + str(i)].value
				else:
					ws["AC" + str(i)] = float(ws["G" + str(i)].value + ws["H" + str(i)].value)/(ws["I" + str(i)].value)

			if ws["T" + str(i)].value != None:

				if ws["T" + str(i)].value == 0:
					ws["AD" + str(i)] = ws["R" + str(i)].value + ws["S" + str(i)].value
				else:
					ws["AD" + str(i)] = float(ws["R" + str(i)].value + ws["S" + str(i)].value)/ws["T" + str(i)].value

			i += 1


	wb.save("processedGamesKDA.xlsx")

def normalizeTest(playerIds, wb, wb_s, wb_t):

	letters = ["C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O", "P", "Q", "R", "S", "T", "U", "V", "W", "X", "Y", "Z", "AA", "AB", "AC", "AD"]
	
	for letter in letters:

		print(letter)

		vals = []

		for playerId in playerIds:

			current_ws = wb[str(playerId)]

			i = 2

			while current_ws["A" + str(i)].value != None:

				if current_ws[letter + str(i)].value != None:

					vals.append(float(current_ws[letter + str(i)].value))

				i += 1

		sigma = np.std(vals)
		mu = np.mean(vals)
		print(mu)

		current_ws = wb_t["39550290"]
		update_ws = wb_s["39550290"]

		i = 2

		while current_ws["A" + str(i)].value != None:

			# when data is missing
			if current_ws[letter + str(i)].value == None:

				update_ws[letter + str(i)] = 0

			else:

				update_ws[letter + str(i)] = (current_ws[letter + str(i)].value - mu)/sigma

			i += 1

	wb_s.save("dataSetKDATest.xlsx")


def normalize(playerIds, wb, wb_s):

	letters = ["C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O", "P", "Q", "R", "S", "T", "U", "V", "W", "X", "Y", "Z", "AA", "AB", "AC", "AD"]
	
	for letter in letters:

		print(letter)

		vals = []

		for playerId in playerIds:

			current_ws = wb[str(playerId)]

			i = 2

			while current_ws["A" + str(i)].value != None:

				if current_ws[letter + str(i)].value != None:

					vals.append(float(current_ws[letter + str(i)].value))

				i += 1

		sigma = np.std(vals)
		mu = np.mean(vals)
		print(mu)

		for playerId in playerIds:

			current_ws = wb[str(playerId)]
			update_ws = wb_s[str(playerId)]

			i = 2

			while current_ws["A" + str(i)].value != None:

				# when data is missing
				if current_ws[letter + str(i)].value == None:

					update_ws[letter + str(i)] = 0

				else:

					update_ws[letter + str(i)] = (current_ws[letter + str(i)].value - mu)/sigma

				i += 1


	wb_s.save("dataSetKDA.xlsx")

		

main()
