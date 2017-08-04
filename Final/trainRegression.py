import csv
from keras.models import Sequential
from keras.layers import Dense, Activation, Dropout, Merge, LSTM, Embedding
from keras.regularizers import l2, activity_l2
import keras
import numpy as np 
from openpyxl import load_workbook, Workbook

import matplotlib.pyplot as plt

def main():

	# fixData()
	# normalizeData()
	# regressionModel()

	regressionModel()
	# generateSimilarSet("normalizedRegression_removed.csv", 2)
	# sequentialRegressionModel()
	# realRegression()
	
def regressionModel():

	X, y = generateDataSet("normalizedRegression_removed.csv")
	
	# print(X)
	# print(y)

	'''
        Playing around with different layers.
        '''
	model = Sequential()
	# model.add(Dropout(0.5, input_shape = (12,)))
	model.add(Dense(128, input_dim = 12, init = 'normal', activation = "relu", W_regularizer = keras.regularizers.WeightRegularizer(l1=0., l2=0.)))
	# model.add(LSTM(32, input_shape(4, 22)))
	model.add(Dropout(0.6))
	model.add(Dense(128, input_dim = 12, init = 'normal', activation = "relu", W_regularizer = keras.regularizers.WeightRegularizer(l1=0., l2=0.)))
	model.add(Dropout(0.6))
	# model.add(Dense(128, init = 'normal', activation = "relu", W_regularizer = keras.regularizers.WeightRegularizer(l1=0., l2=0.)))
	# model.add(Dropout(0.6))
	# model.add(Dropout(0.8))
	# model.add(Dense(128, activation = "relu"))
	# model.add(Dropout(0.2))
	# model.add(Dense(128, activation = "relu"))
	# model.add(Dropout(0.2))
	# model.add(Dense(64, activation = "relu"))
	# model.add(Dropout(0.5))
	model.add(Dense(10, init = 'normal', activation = "linear"))

	model.compile(loss='mse',
              optimizer=keras.optimizers.RMSprop(lr=0.001, rho=0.9, epsilon=1e-08, decay=0.0))

	history = model.fit(X, y, validation_split = 0.1,nb_epoch = 1000, batch_size = len(X))

	plt.plot(history.history['loss'])
	plt.plot(history.history['val_loss'])
	plt.title('Regression MSE')
	plt.ylabel('MSE')
	plt.xlabel('Epoch')
	plt.legend(['Train', 'Validation'], loc='upper left')
	plt.show()

	# for layer in model.layers:

	# 	print layer.get_weights()

	# print(model.predict(X))
	# print(y)

def sequentialRegressionModel():

	time_window = 8
	X, y = generateSimilarSet("normalizedRegression_removed.csv", time_window)
	model = Sequential()
	model.add(Dense(128, input_dim = time_window, activation = "relu", W_regularizer = keras.regularizers.WeightRegularizer(l1=0., l2=0.) ))
	model.add(Dropout(0.8))
	model.add(Dense(128, activation="relu",  W_regularizer = keras.regularizers.WeightRegularizer(l1=0., l2=0.)))
	model.add(Dropout(0.8))
	model.add(Dense(1, activation = "linear"))

	model.compile(loss='mse',
              optimizer='rmsprop')

	model.fit(X, y, validation_split = 0.1, nb_epoch = 10000, batch_size = 106-time_window)


def realRegression():

	X, y = generateDataSet("merged.csv")

	model = Sequential()

	model.add(Dense(128, input_dim = 10, activation = "relu"))
	model.add(Dropout(0.4))
	model.add(Dense(128, activation = "relu"))
	model.add(Dropout(0.4))
	model.add(Dense(10, activation = "linear"))

	model.compile(loss='mse',
              optimizer='adadelta')

	history = model.fit(X, y, validation_split = 0.1,nb_epoch = 1000, batch_size = len(X))

	plt.plot(history.history['loss'])
	plt.plot(history.history['val_loss'])
	plt.title('model MSE')
	plt.ylabel('MSE')
	plt.xlabel('epoch')
	plt.legend(['train', 'validation'], loc='upper left')
	plt.show()
			


def fixData():

	players = [66271229, 44207669, 47063254, 56723622, 34399881, 47837396, 19305039, 37348109, 51635691, 28360391]

	wb = load_workbook("processedGamesKDA_new1.xlsx")

	letters = ["C", "D", "E", "F", "H", "X", "M", "N", "O", "Q", "V", "W"]

	for i in range(len(players)):

		ws = wb[str(players[i])]

		for letter in letters:

			index = 2

			values = []

			while ws["A" + str(index)].value != None:

				if ws[letter + str(index)].value != None:

					values.append(ws[letter + str(index)].value)

				index += 1

			player_mean = np.mean(values)

			index = 2

			while ws["A" + str(index)].value != None:

				if ws[letter + str(index)].value == None or ws[letter + str(index)].value - player_mean < .0001:

					ws[letter + str(index)] = None

				index += 1
		
	# wb.save("processedGamesKDA_new_playerMean.xlsx")

	wb.save("processedGamesKDA_new2.xlsx")

def normalizeData():

	wb = load_workbook("processedGamesKDA_new.xlsx")

	new_wb = load_workbook("normalizedRegression_removed.xlsx")
	new_ws = new_wb["training"]


	# input_letters = ["C", "D", "E", "F", "H", "X", "M", "N", "O", "Q", "V", "W"]
	letters = ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O", "P", "Q", "R", "S", "T", "U", "V", "W", "X"]
	letter_data = {}

	players = [66271229, 44207669, 47063254, 56723622, 34399881, 47837396, 19305039, 37348109, 51635691, 28360391]


	for letter in letters:

		if letter != "A" or letter != "B":

			vals = []

			for playerId in players:

				ws = wb[str(playerId)]

				index = 2

				while ws["A" + str(index)].value != None:

					if ws[letter + str(index)].value != None:

						vals.append(ws[letter + str(index)].value)

					index += 1

			mu = np.mean(vals)
			sigma = np.std(vals)

			letter_data[letter] = (mu,sigma)

	i = 2
	for playerId in players:

		ws = wb[str(playerId)]

		index = 2

		while ws["A" + str(index)].value != None:

			if ws["C" + str(index)].value == None or ws["H" + str(index)].value == None or ws["M" + str(index)].value == None or ws["R" + str(index)].value == None:

				index += 1
				continue

			for letter in letters:

				mu = letter_data[letter][0]
				sigma = letter_data[letter][1]

				if letter == "A" or letter == "B":

					new_ws[letter + str(i)].value = ws[letter + str(index)].value

				else:

					new_ws[letter + str(i)].value = float(ws[letter + str(index)].value-mu)/sigma

			index += 1
			i += 1



	new_wb.save("normalizedRegression_removed.xlsx")

def generateSimilarSet(file_name, time_window):

	X = []
	y = []

	# input_ranges = [(2,6), (7,8), (12,16), (17,18), (22,24)]
	# input_ranges = [(2,6), (7,8), (23,24)]
	input_ranges = [(8,9)]
	output_ranges = [(8,9)]

	i = 0

	player_rows = [[],[],[],[],[],[],[],[],[],[]]
	with open(file_name, 'rb') as csvfile:
		reader = csv.reader(csvfile, delimiter = ",", quotechar = "|")
		current_player = 0
		previousMatchId = 0
		for row in reader:
			if i == 0:
				i+=1
				continue

			matchId = row[0]

			if previousMatchId == 0:

				previousMatchId = row[0]
			

			if matchId<previousMatchId:

				current_player += 1
				# print(float(matchId))
			# print(current_player)
			player_rows[current_player].append(row)
			previousMatchId = matchId

	for i in range(time_window, 106):

		for j in range(len(player_rows)):

			current_x = []
			current_y = []

			for k in range(i-time_window, i):

				for l in range(len(input_ranges)):

					bounds = input_ranges[l]

					for m in range(bounds[0], bounds[1]):

						current_x.append(float(player_rows[j][k][m]))

			for k in range(len(output_ranges)):

				bounds = output_ranges[k]

				for l in range(bounds[0], bounds[1]):

					current_y.append(float(player_rows[j][i][l]))

			X.append(current_x)
			y.append(current_y)

	X = np.asarray(X)
	y = np.asarray(y)
	print(X)
	print(y)

	return X, y


def generateDataLSTM(file_name, time_window):

	X = []
	y = []

	input_ranges = [(2,6), (7,8), (12,16), (17,18), (22,24)]
	# input_ranges = [(2,6), (7,8), (23,24)]
	output_ranges = [(8,10)]

	i = 0
	with open(file_name, 'rb') as csvfile:
		reader = csv.reader(csvfile, delimiter = ",", quotechar = "|")
		for row in reader:
			if i==0:
				i += 1
				# print(len(row[2:]))
				continue
			if i == 256:
				break
			current_x = []
			current_y = []

			for j in range(len(input_ranges)):

				bounds = input_ranges[j]

				for k in range(bounds[0], bounds[1]):

					current_x.append(float(row[k]))

			for l in range(len(output_ranges)):

				bounds = output_ranges[l]

				for m in range(bounds[0], bounds[1]):

					current_y.append(float(row[m]))

			X.append(current_x)
			y.append(current_y)
			i+=1

	realX = []
	for i in range(time_window):

		current = []

		for j in range(i-time_window, i):

			current.append(X[j])

		realX.append(current)

	realX = np.asarray(realX)

	return realX, y[time_window:]

def generateDataSet(file_name):

	X = []
	y = []

	input_ranges = [(2,6), (7,8), (12, 16), (17,18), (22,24)]
	output_ranges = [(6,7), (8,12), (16,17), (18, 22)]
	# input_ranges = [(1,11)]
	# input_ranges = [(2,6), (7,8), (23,24)]
	# output_ranges = [(11,21)]

	i = 0
	with open(file_name, 'rb') as csvfile:
		reader = csv.reader(csvfile, delimiter = ",", quotechar = "|")
		for row in reader:
			if i == 0:
				i += 1
				continue
			current_x = []
			current_y = []

			for j in range(len(input_ranges)):

				bounds = input_ranges[j]

				for k in range(bounds[0], bounds[1]):

					current_x.append(float(row[k]))

			for l in range(len(output_ranges)):

				bounds = output_ranges[l]

				for m in range(bounds[0], bounds[1]):

					current_y.append(float(row[m]))

			X.append(current_x)
			y.append(current_y)
			i+=1

	X = np.asarray(X)
	y = np.asarray(y)

	# print(X)
	# print(y)

	return X, y
			


main()
