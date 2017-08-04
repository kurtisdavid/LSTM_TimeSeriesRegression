import requests
from collections import deque
import simplejson as json
import glob
import time
import sys
import os
from openpyxl import load_workbook, Workbook

# average rates are already calculated in processedGames, so just update the values not to be averaged!

def main():

	wb = load_workbook("processedGamesKDA_new.xlsx")
	wb1 = load_workbook("processedGamesKDA_new1.xlsx")
	# players = [66271229, 44207669, 47063254, 56723622, 34399881, 47837396, 19305039, 37348109, 51635691, 28360391]
	for i in range(len(players)):

		print(players[i])
		# wb = fixKDA(players[i], wb)
		wb1 = movingAverage(players[i], wb, wb1)
		# break
		# wb = collectStats(players[i], wb)
		# return

def movingAverage(playerId, wb, wb1):
	
	
	ws1 = wb1[str(playerId)]
	ws = wb[str(playerId)]

	mapping_key = {"fb": "H", "win": "C", "n_games": "X"}

	i = 2

	while ws["A" + str(i)].value != None:

		
		champion = ws["F" + str(i)].value

		p_stats = {"fb": 0, "win": 0, "n_games": 0}

		count = 0

		for j in range(i-1, 2, -1):

			# print(j)

			# only get past 25 games of specific champion
			if count == 25:

				break

			if str(ws["F" + str(j)].value) != str(champion):

				continue

			matchId = ws["A" + str(j)].value
			directory = "./" + str(playerId) + "/" + str(matchId) + ".json"

			current = {}

			with open(directory) as data_file:

				data_json = json.load(data_file)
				current = innerStats(data_json, playerId)

			# print(current)
			for key in current:

				p_stats[key] += current[key]

			p_stats["n_games"] += 1
			count += 1

		# print(p_stats)
		if p_stats["n_games"] == 0:

			for key in mapping_key:

				if key == "n_games":

					ws1[mapping_key[key] + str(i)].value = 0

				else:

					ws1[mapping_key[key] + str(i)].value = None

		else:

			for key in p_stats:

				if key == "n_games":
					ws1[mapping_key[key] + str(i)] = p_stats["n_games"]
					continue

				p_stats[key] = float(p_stats[key])/p_stats["n_games"]
				ws1[mapping_key[key] + str(i)] = p_stats[key]

		i += 1
	wb1.save("processedGamesKDA_new1.xlsx")
	return wb1


def innerStats(matchJSON, primaryId):

	stats = {}

	# remade game, not valid
	if 180 <= matchJSON["matchDuration"] <= 300:

		return stats

	if matchJSON["matchDuration"] <= 1200:

		return stats

	pID = -1

	for identity in matchJSON["participantIdentities"]:

		if str(identity["player"]["summonerId"])==str(primaryId):

			pID = identity["participantId"]

	for participant in matchJSON["participants"]:

		if participant["participantId"] == pID:

			p_stats = participant["stats"]

			if str(p_stats["firstBloodAssist"]) == "True" or str(p_stats["firstBloodKill"]) == "True":
				
				stats["fb"] = 1

			else:
				stats["fb"] = 0

			if str(p_stats["winner"]) == "True":

				stats["win"] = 1

			else:

				stats["win"] = 0

	return stats

def fixKDA(playerId, wb):

	ws = wb[str(playerId)]
	i = 2

	while ws["A" + str(i)].value != None:

		print(i)

		if ws["I" + str(i)].value == 0:

			ws["AC" + str(i)] = ws["G" + str(i)].value + ws["H" + str(i)].value

		else:

			ws["AC" + str(i)] = float(ws["G" + str(i)].value + ws["H" + str(i)].value)/ws["I" + str(i)].value

		if ws["T" + str(i)].value == 0:

			ws["AD" + str(i)] = ws["R" + str(i)].value + ws["S" + str(i)].value

		else:

			ws["AD" + str(i)] = float(ws["R" + str(i)].value + ws["S" + str(i)].value)/ws["T" + str(i)].value

		i += 1

	wb.save("processedGamesKDA_new.xlsx")
	return wb


def collectStats(playerId, wb):

	mapping_key = {"p1": {"kills": "G", "assists": "H", "deaths": "I", "cs10": "K", "cs20": "L", "xp10": "M", "xp20": "N"},
					"p2": {"kills": "R", "assists": "S", "deaths": "T", "cs10": "V", "cs20": "W", "xp10": "X", "xp20": "Y"}}

	ws = wb[str(playerId)]

	i = 2

	while ws["A" + str(i)].value != None:

		matchId = str(ws["A" + str(i)].value)

		file_name = "./" + str(playerId) + "/" + matchId + ".json"

		stats = None

		with open(file_name) as data_file:

			data = json.load(data_file)
			stats = updateGameStats(data, playerId)

		if i==92:
			print(stats)
		if stats["p2"] == {}:

			i+=1
			continue

		for key in stats:

			for sub_key in stats[key]:

				ws[mapping_key[key][sub_key] + str(i)] = stats[key][sub_key]

		i += 1

	wb.save("processedGamesKDA_new.xlsx")
	return wb



def updateGameStats(matchJSON, primaryId):

	stats = {"p1": {}, "p2": {}}
	file = open("checkAfterNew.txt", "a")

	# remade game, not valid
	if 180 <= matchJSON["matchDuration"] <= 300:

		return stats

	if matchJSON["matchDuration"] < 1200:

		file.write(str(primaryId) + " - " + str(matchJSON["matchId"]) + "\n")
		return stats


	game_ids = []

	for participant in matchJSON["participants"]:

			if participant["timeline"]["lane"] == "TOP":

				game_ids.append(participant["participantId"])

	# Riot API did not recognize the right roles, invalid game to be used
	if len(game_ids) == 1:

		return stats

	pID_1 = 0
	pID_2 = 0

	for identity in matchJSON["participantIdentities"]:

		if identity["participantId"] in game_ids and str(identity["player"]["summonerId"])!=str(primaryId):

			pID_2 = identity["participantId"]
		
		elif identity["participantId"] in game_ids:

			pID_1 = identity["participantId"]
	
	for participant in matchJSON["participants"]:

		if participant["participantId"] == pID_1:

			p_stats = participant["stats"]
			p_timeline = participant["timeline"]

			stats["p1"]["kills"] = p_stats["kills"]
			stats["p1"]["assists"] = p_stats["assists"]
			stats["p1"]["deaths"] = p_stats["deaths"]

			stats["p1"]["cs10"] = p_timeline["creepsPerMinDeltas"]["zeroToTen"]
			stats["p1"]["cs20"] = p_timeline["creepsPerMinDeltas"]["tenToTwenty"]

			stats["p1"]["xp10"] = p_timeline["xpPerMinDeltas"]["zeroToTen"]
			stats["p1"]["xp20"] = p_timeline["xpPerMinDeltas"]["tenToTwenty"]


		if participant["participantId"] == pID_2:

			p_stats = participant["stats"]
			p_timeline = participant["timeline"]

			stats["p2"]["kills"] = p_stats["kills"]
			stats["p2"]["assists"] = p_stats["assists"]
			stats["p2"]["deaths"] = p_stats["deaths"]

			stats["p2"]["cs10"] = p_timeline["creepsPerMinDeltas"]["zeroToTen"]
			stats["p2"]["cs20"] = p_timeline["creepsPerMinDeltas"]["tenToTwenty"]

			stats["p2"]["xp10"] = p_timeline["xpPerMinDeltas"]["zeroToTen"]
			stats["p2"]["xp20"] = p_timeline["xpPerMinDeltas"]["tenToTwenty"]

	return stats




main()